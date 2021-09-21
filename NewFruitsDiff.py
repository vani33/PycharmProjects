import openpyxl, xlsxwriter
from openpyxl import load_workbook

workbook = openpyxl.load_workbook("C:\\Users\\kothapalli.vanisree\\OneDrive - HCL Technologies Ltd\\Desktop\\FruitsDifference.xlsx") #To load the workbook
sh1 = workbook['Sheet1'] #Current sheet in workbook

# appleDay1 =sh1['B2'].value
# #print(appleDay1)
# appleDay2 =sh1['C2'].value
# #print(appleDay2)
# appleDif = abs(appleDay1 - appleDay2)
# #print(appleDif)
#
# appDifPercent = (abs(appleDay1 - appleDay2)/appleDay1)*100
# #print(appDifPercent)

rows = sh1.max_row
columns = sh1.max_column - 2
#print(rows) #To print the no. of rows
#print(columns)

list1 = []
list2 = []
for i in range(2,rows + 1):
    for j in range(2, columns):
       list1.append(sh1.cell(i,j).value)
print(list1)

for i in range(2,rows + 1):
    for j in range(3, columns + 1):
       list2.append(sh1.cell(i,j).value)
print(list2)

diff = []
diffPercent = []
zip_object = zip(list1, list2)
for list1_i, list2_i in zip_object:
    diff.append(abs(list1_i - list2_i))
    diffPercent.append(((abs(list1_i - list2_i))/list1_i)*100)

print(diff) #To print the difference between day1 and day2
print(diffPercent)#To print the percentage of difference


file = "C:\\Users\\kothapalli.vanisree\\OneDrive - HCL Technologies Ltd\\Desktop\\FruitsDifference.xlsx"
wb = openpyxl.load_workbook(filename=file)
ws = wb['Sheet1']
col = ws.max_column - 1

for item1 in range(2,len(diff)+2):
    print(ws.cell(row=item1, column=col, value=diff[item1 - 2]))
for item2 in range(2,len(diffPercent)+2):
    print(ws.cell(row=item2, column=col + 1, value=diffPercent[item2 - 2]))

wb.save(file)



