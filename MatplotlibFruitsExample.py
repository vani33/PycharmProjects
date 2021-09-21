import openpyxl
from matplotlib import pyplot as plt
import pandas as pd

workbook = openpyxl.load_workbook(
    "C:\\Users\\kothapalli.vanisree\\OneDrive - HCL Technologies Ltd\\Desktop\\MatplotlibFruitsExample.xlsx")  # To load the workbook
sh1 = workbook['Sheet1']  # Current sheet in workbook

rows = sh1.max_row
columns = sh1.max_column - 4
# print(rows)
# print(columns)

day1 = []
day2 = []
day3 = []
for i in range(2, rows + 1):
    for j in range(2, columns - 1):
        day1.append(sh1.cell(i, j).value)
print(day1)

for i in range(2, rows + 1):
    for j in range(3, columns):
        day2.append(sh1.cell(i, j).value)
print(day2)

for i in range(2, rows + 1):
    for j in range(4, columns + 1):
        day3.append(sh1.cell(i, j).value)
print(day3)

diff1and2 = []
diff2and3 = []
diffPercent1and2 = []
diffPercent2and3 = []
zip_object1and2 = zip(day1, day2)
zip_object1and3 = zip(day2, day3)
for day1_i, day2_i in zip_object1and2:
    diff1and2.append(round(day1_i - day2_i))
    diffPercent1and2.append(round((abs(day1_i-day2_i)/day1_i) * 100))

for day2_i, day3_i in zip_object1and3:
    diff2and3.append(round(day2_i - day3_i))
    diffPercent2and3.append(round((abs(day2_i-day3_i)/day2_i) * 100))

print(diff1and2)
print(diff2and3)
print(diffPercent1and2)
print(diffPercent2and3)

file = "C:\\Users\\kothapalli.vanisree\\OneDrive - HCL Technologies Ltd\\Desktop\\MatplotlibFruitsExample.xlsx"
wb = openpyxl.load_workbook(filename=file)
ws = wb['Sheet1']
col = ws.max_column

for item1 in range(2, rows + 1):
    print(ws.cell(row=item1, column=col-3, value=diffPercent1and2[item1 - 2]))
for item2 in range(2, rows + 1):
    print(ws.cell(row=item2, column=col-2, value=(diffPercent2and3[item2 - 2])))
for item3 in range(2, rows + 1):
    print(ws.cell(row=item3, column=col-1, value=diff1and2[item3 - 2]))
for item4 in range(2, rows + 1):
    print(ws.cell(row=item4, column=col, value=(diff2and3[item4 - 2])))

wb.save(file)

#####################################################################################################


#Line graph
# data = []
# for i in range(2, sh1.max_row + 1):
#     d = []
#     for j in range(1, sh1.max_column - 1):
#         d.append(sh1.cell(i, j).value)
#     data.append(d)
# print(type(data))
# df1 = pd.DataFrame(data, columns=["Fruits", "Day1", "Day2", "Day3"])
# df1.plot(x="Fruits", y=["Day1", "Day2", "Day3"])
# plt.xlabel("Name of Fruits",fontname="Times New Roman",fontweight="bold",fontsize=15)
# plt.ylabel("Price range(in Rupees)",fontname="Times New Roman",fontweight="bold",fontsize=15)
# plt.yticks(range(100,300,20))
# plt.title("Price of Fruits daywise",fontname="Times New Roman",fontweight="bold", fontsize=22)
# plt.show()


#Line graph in simple way
# df = pd.read_excel(
#     'C:\\Users\\kothapalli.vanisree\\OneDrive - HCL Technologies Ltd\\Desktop\\MatplotlibFruitsExample.xlsx', 'Sheet1')
# df.plot(x="Fruits", y=["Day1","Day2","Day3"])
# plt.xlabel("Name of Fruits",fontname="Times New Roman",fontweight="bold",fontsize=15)
# plt.ylabel("Price range(in Rupees)",fontname="Times New Roman",fontweight="bold",fontsize=15)
# #plt.yticks(range(0,300,20))
# plt.title("Price of Fruits daywise",fontname="Times New Roman",fontweight="bold",fontsize=20)
# plt.grid()
#
# #Plot of difference%
# df.plot(x="Fruits", y=["Difference%(Day1-Day2)","Difference%(Day2-Day3)"])
# plt.xlabel("Name of Fruits",fontname="Times New Roman",fontweight="bold",fontsize=15)
# plt.ylabel("Difference Percentage(in %)",fontname="Times New Roman",fontweight="bold",fontsize=15)
# #plt.yticks(range(0,300,20))
# plt.title("Difference Percentage of Fruits daywise",fontname="Times New Roman",fontweight="bold",fontsize=20)
# plt.grid()
#
# #plot of difference
# df.plot(x="Fruits", y=["Difference(Day1-Day2)","Difference(Day2-Day3)"])
# plt.xlabel("Name of Fruits",fontname="Times New Roman",fontweight="bold",fontsize=15)
# plt.ylabel("Price Difference(in Rupees)",fontname="Times New Roman",fontweight="bold",fontsize=15)
# #plt.yticks(range(0,300,20))
# plt.title("Price Difference of Fruits daywise",fontname="Times New Roman",fontweight="bold",fontsize=20)
# plt.grid()
#
#
# plt.show()
####################################################################

#Bar Graph
# data = [["Apple",100,200,150],
#         ["Mango",150,220,280],
#         ["PineApple",155,225,125],
#         ["Kiwi",135,15,200]]

# data = []
# for i in range(2, sh1.max_row + 1):
#     d = []
#     for j in range(1, sh1.max_column - 1):
#         d.append(sh1.cell(i, j).value)
#     data.append(d)
# print(data)
#
# df1 = pd.DataFrame(data, columns=["Fruits", "Day1", "Day2", "Day3"])
# print(df1)
#
# df1.plot(x="Fruits", y=["Day1", "Day2", "Day3"], kind="bar")
# plt.xlabel("Name of Fruits",fontname="Times New Roman",fontweight="bold",fontsize=15)
# plt.ylabel("Price range(in Rupees)",fontname="Times New Roman",fontweight="bold",fontsize=15)
# plt.yticks(range(100,300,20))
# plt.title("Price of Fruits daywise",fontname="Times New Roman",fontweight="bold",fontsize=20)
# plt.show()






# #Bar graph in simple way
df = pd.read_excel(
    'C:\\Users\\kothapalli.vanisree\\OneDrive - HCL Technologies Ltd\\Desktop\\MatplotlibFruitsExample.xlsx', 'Sheet1')

df.plot(x="Fruits", y=["Day1","Day2","Day3"], kind="bar")
plt.xlabel("Name of Fruits",fontname="Times New Roman",fontweight="bold",fontsize=15)
plt.ylabel("Price range(in Rupees)",fontname="Times New Roman",fontweight="bold",fontsize=15)
plt.yticks(range(0,300,20))
plt.title("Price of Fruits daywise",fontname="Times New Roman",fontweight="bold",fontsize=20)
plt.legend()
plt.grid()

#Plot of difference%
df.plot(x="Fruits", y=["Difference%(Day1-Day2)","Difference%(Day2-Day3)"], kind="bar")
#plt.scatter(x="Fruits", y=["Difference%(Day1-Day2)","Difference%(Day2-Day3)"], kind="bar",marker="*")
plt.xlabel("Name of Fruits",fontname="Times New Roman",fontweight="bold",fontsize=15)
plt.ylabel("Difference Percentage(in %)",fontname="Times New Roman",fontweight="bold",fontsize=15)
#plt.yticks(range(0,300,20))
plt.title("Difference Percentage of Fruits daywise",fontname="Times New Roman",fontweight="bold",fontsize=20)
plt.grid()

#plot of difference
df.plot(x="Fruits", y=["Difference(Day1-Day2)","Difference(Day2-Day3)"], kind="bar")
plt.axhline(color='black')
plt.xlabel("Name of Fruits",fontname="Times New Roman",fontweight="bold",fontsize=15)
plt.ylabel("Price Difference(in Rupees)",fontname="Times New Roman",fontweight="bold",fontsize=15)
#plt.yticks(range(0,300,20))
plt.title("Price Difference of Fruits daywise",fontname="Times New Roman",fontweight="bold",fontsize=20)
plt.grid()
plt.show()

########################################################################################################
